# data_fetching_system/data_fetcher.py

import os
import json
import requests
import pandas as pd
import mplfinance as mpf
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from datetime import datetime, timedelta
from dotenv import load_dotenv
from analyzer.symbol_lookup import resolve_symbol

load_dotenv()

# === API ENDPOINTS
HISTORICAL_URL = "https://apiconnect.angelone.in/rest/secure/angelbroking/historical/v1/getCandleData"
OI_URL = "https://apiconnect.angelone.in/rest/secure/angelbroking/historical/v1/getOIData"

# === HEADERS
HEADERS = {
    "X-UserType": "USER",
    "X-SourceID": "WEB",
    "X-ClientLocalIP": os.getenv("LOCAL_IP"),
    "X-ClientPublicIP": os.getenv("PUBLIC_IP"),
    "X-MACAddress": os.getenv("MAC_ADDRESS"),
    "X-PrivateKey": os.getenv("ANGEL_API_KEY"),
    "Authorization": f"Bearer {os.getenv('ANGEL_AUTH_TOKEN')}",
    "Accept": "application/json",
    "Content-Type": "application/json"
}

# === INTERVAL LIMITS
INTERVAL_LIMITS = {
    "ONE_MINUTE": 30,
    "THREE_MINUTE": 60,
    "FIVE_MINUTE": 100,
    "TEN_MINUTE": 100,
    "FIFTEEN_MINUTE": 200,
    "THIRTY_MINUTE": 200,
    "ONE_HOUR": 400,
    "ONE_DAY": 2000,
}

def get_time_range(interval: str, days: int):
    now = datetime.now()
    if interval in ["ONE_MINUTE", "THREE_MINUTE", "FIVE_MINUTE", "TEN_MINUTE", "FIFTEEN_MINUTE", "THIRTY_MINUTE"]:
        to_date = now.replace(hour=15, minute=15, second=0, microsecond=0)
        from_date = to_date - timedelta(days=days)
        from_date = from_date.replace(hour=9, minute=15)
    else:
        to_date = now
        from_date = to_date - timedelta(days=days)
    return from_date, to_date

def enrich_ohlc_df(df: pd.DataFrame) -> pd.DataFrame:
    df["datetime"] = pd.to_datetime(df["datetime"]).dt.tz_localize(None)
    df["price_change"] = df["close"] - df["open"]
    df["ltp_change_%"] = df["close"].pct_change().fillna(0) * 100
    df["avg_price"] = (df["high"] + df["low"]) / 2
    df["candle_range"] = df["high"] - df["low"]
    df["date"] = df["datetime"].dt.strftime("%d-%m-%Y")
    df["time"] = df["datetime"].dt.strftime("%H:%M")
    return df

def fetch_historical_candles(symbol: str, interval: str, days: int) -> pd.DataFrame:
    interval = interval.upper()
    if interval not in INTERVAL_LIMITS or days > INTERVAL_LIMITS[interval]:
        raise ValueError(f"❌ Invalid interval '{interval}' or days limit exceeded.")

    resolved = resolve_symbol(symbol)
    start, end = get_time_range(interval, days)

    payload = {
        "exchange": resolved["exchange"],
        "symboltoken": resolved["token"],
        "interval": interval,
        "fromdate": start.strftime("%Y-%m-%d %H:%M"),
        "todate": end.strftime("%Y-%m-%d %H:%M")
    }

    try:
        print(f"📡 Fetching historical data for {symbol} [{interval}] ({days}d)")
        res = requests.post(HISTORICAL_URL, headers=HEADERS, json=payload)
        res.raise_for_status()
        candles = res.json().get("data", [])
        if not candles:
            print("⚠️ No candle data returned.")
            return pd.DataFrame()

        df = pd.DataFrame(candles, columns=["datetime", "open", "high", "low", "close", "volume"])
        return enrich_ohlc_df(df)

    except Exception as e:
        print(f"❌ Historical fetch failed for {symbol}: {e}")
        return pd.DataFrame()

def fetch_historical_oi(symbol: str, interval: str, days: int) -> pd.DataFrame:
    interval = interval.upper()
    if interval not in INTERVAL_LIMITS or days > INTERVAL_LIMITS[interval]:
        raise ValueError("❌ Invalid interval or days limit.")

    resolved = resolve_symbol(symbol)
    start, end = get_time_range(interval, days)

    payload = {
        "exchange": resolved["exchange"],
        "symboltoken": resolved["token"],
        "interval": interval,
        "fromdate": start.strftime("%Y-%m-%d %H:%M"),
        "todate": end.strftime("%Y-%m-%d %H:%M")
    }

    try:
        print(f"📡 Fetching OI data for {symbol}")
        res = requests.post(OI_URL, headers=HEADERS, json=payload)
        res.raise_for_status()
        oi_data = res.json().get("data", [])
        if not oi_data:
            print("⚠️ No OI data returned.")
            return pd.DataFrame()

        df = pd.DataFrame(oi_data)
        df["time"] = pd.to_datetime(df["time"]).dt.tz_localize(None)
        return df

    except Exception as e:
        print(f"❌ OI data fetch failed: {e}")
        return pd.DataFrame()

def plot_monthly_candlesticks(df: pd.DataFrame, symbol: str):
    if df.empty:
        print("⚠️ No data to plot.")
        return

    df["datetime"] = pd.to_datetime(df["date"] + " " + df["time"], dayfirst=True)
    df.set_index("datetime", inplace=True)
    df["month"] = df.index.to_period("M")

    os.makedirs("chart", exist_ok=True)

    for month in df["month"].unique():
        df_month = df[df["month"] == month]
        df_candle = df_month[["open", "high", "low", "close", "volume"]]

        mpf.plot(
            df_candle,
            type="candle",
            style="charles",
            title=f"{symbol.upper()} | {month}",
            ylabel="Price",
            volume=True,
            mav=(5, 10),
            figratio=(14, 7),
            tight_layout=True,
            savefig=f"chart/{symbol.upper()}-{month}.png"
        )
        print(f"✅ Saved chart: chart/{symbol.upper()}-{month}.png")

def save_to_excel_with_chart(df: pd.DataFrame, symbol: str):
    if df.empty:
        print("⚠️ No data to save.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "OHLC Analysis"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)

    chart = LineChart()
    chart.title = f"{symbol.upper()} Closing Price Trend"
    chart.y_axis.title = "Price"
    chart.x_axis.title = "Candle #"

    max_row = ws.max_row
    price_col = df.columns.get_loc("close") + 1
    chart_data = Reference(ws, min_col=price_col, min_row=1, max_row=max_row)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=max_row))  # x-axis from datetime
    ws.add_chart(chart, "K2")

    filename = f"{symbol.upper()}_OHLC_REPORT.xlsx"
    save_path = os.path.join("chart", filename)
    wb.save(save_path)
    print(f"✅ Saved Excel report: {save_path}")
